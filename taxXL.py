import openpyxl
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter

class xlwbook:
    def __init__(xlw):
        xlw.f_name = ""
        xlw.w_dict = {}
        xlw.month_sel = ""
    def openbook (xlw):
        lst_amt = [xlw.month_sel]
        for x, y in xlw.w_dict.items():
            if y == "":
                y = "+0"
            y = y.replace(",", "")
            lst_amt.append(y)
        wb = load_workbook(xlw.f_name)
        ws = wb.active
        for row in range(6, 18):
            lst_i = 0
            lst_v = ""
            for col in range(1, 28):
                char = get_column_letter(col) + str(row)
                if col == 1:
                   s_month = ws[char].value
                if (s_month.lower() == xlw.month_sel) and get_column_letter(col) != "A":
                    lst_var = lst_amt[lst_i]
                    if ws[char].value == None:
                        lst_v = lst_var.replace("+", "=", 1)
                        ws[char].value = lst_v
                        #print(ws[char].value)
                    else:
                        if lst_var != "+0":
                            lst_v = ws[char].value
                            lst_v = lst_v + lst_var
                            ws[char].value = lst_v
                        #print(ws[char].value)
                lst_v = ""
                lst_i += 1
        wb.save(xlw.f_name)
        wb.close()
